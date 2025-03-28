from flask import Flask, render_template, jsonify, request, send_from_directory, session
import cv2
import subprocess
import openpyxl
import firebase_admin
from firebase_admin import credentials, db
import os
from qr import scan_qr_code
from testing import query_gemini
from dotenv import load_dotenv

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

load_dotenv()

FIREBASE_JSON = os.getenv('FIREBASE_JSON')  
FIREBASE_URL = os.getenv('FIREBASE_URL')

cred = credentials.Certificate(FIREBASE_JSON)
firebase_admin.initialize_app(cred, {
    'databaseURL': FIREBASE_URL
})

EXCEL_FILE_PATH = "datas/heart_rate_data.xlsx"

def generate_frames():
    cap = cv2.VideoCapture(0)
    while True:
        success, frame = cap.read()
        if not success:
            break

        _, buffer = cv2.imencode('.jpg', frame)
        frame = buffer.tobytes()

        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

@app.route('/assets/<path:filename>')
def assets(filename):
    return send_from_directory('assets', filename)

@app.route('/')
def index():
    return render_template('authenticate.html')

@app.route('/scan_qr_code', methods=['POST'])
def scan_qr_code_endpoint():
    scanned_data = scan_qr_code()
    if scanned_data:
        name = scanned_data.get("name", "N/A")
        address = scanned_data.get("address", "N/A")
        gender = "Male" if scanned_data.get("gender") == "M" else "Female"
        dob = scanned_data.get("dob", "N/A")
        
        session['current_mobile'] = scanned_data.get("mobile", "N/A")  

        ref = db.reference("/users")
        ref.child(session['current_mobile']).set({
            "name": name,
            "address": address,
            "gender": gender,
            "dob": dob,
            "mobile": session['current_mobile']
        })
        return jsonify(message="Data successfully sent to Firebase")
    else:
        return jsonify(error="No QR code data found"), 400

@app.route('/start_monitoring', methods=['POST'])
def start_monitoring():
    subprocess.call(["python", "sbm.py"])
    return jsonify(status="Monitoring started")

@app.route('/get_average_bpm', methods=['GET'])
def get_average_bpm():
    current_mobile = session.get('current_mobile')

    if not os.path.exists(EXCEL_FILE_PATH):
        return jsonify(error="BPM data not available"), 404

    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    ws = wb.active

    bpm_values = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1) if ws.cell(row=i, column=2).value]

    if not bpm_values:
        return jsonify(error="No BPM data recorded"), 404

    average_bpm = sum(bpm_values) / len(bpm_values)
    wb.close()

    if current_mobile:
        user_ref = db.reference("/users")
        user_ref.child(current_mobile).update({
            'bpm': average_bpm
        })
    
    return jsonify(average_bpm=average_bpm)

@app.route('/submit_patient_details', methods=['POST'])
def submit_patient_details():
    current_mobile = session.get('current_mobile')
    if not current_mobile:
        return jsonify(error="No user is authenticated"), 400

    data = request.get_json()
    height = data.get('height')
    weight = data.get('weight')
    temperature = data.get('temperature')

    if not (height and weight and temperature):
        return jsonify(error="Incomplete patient details"), 400

    user_ref = db.reference("/users")
    user_ref.child(current_mobile).update({
        "height": height,
        "weight": weight,
        "temperature": temperature
    })
    
    return jsonify(message="Patient details updated successfully")

@app.route('/symptom_input')
def symptom_input():
    return render_template('symptom_input.html')

@app.route('/submit_symptoms', methods=['POST'])
def submit_symptoms():
    current_mobile = session.get('current_mobile')
    if not current_mobile:
        return jsonify(error="No user is authenticated"), 400
    
    symptoms = request.form.get("symptoms")
    
    user_ref = db.reference(f"/users/{current_mobile}")
    user_data = user_ref.get()
    
    if user_data:
        bpm = user_data.get("bpm", "Not Available")
        temperature = user_data.get("temperature", "Not Available")
        gender = user_data.get("gender", "Not Available")
        height = user_data.get("height", "Not Available")
        weight = user_data.get("weight", "Not Available")
    else:
        return jsonify(error="Patient data not available"), 404

    prompt = f"""
    Patient presents with the following:
    - Symptoms: {symptoms}
    - Heart Rate (BPM): {bpm}
    - Temperature: {temperature}°F
    - Gender: {gender}
    - Height: {height}
    - Weight: {weight}
    """

    response = query_gemini(prompt)

    db.reference(f"/users/{current_mobile}/symptoms").push({
        'symptoms': symptoms,
        'AI Suggestion': response,
    })

    return jsonify({"recommended_treatment": response})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
